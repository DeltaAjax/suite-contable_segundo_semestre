"""
amortizacion.py — Módulo de Cálculo de Amortización NIF C-19
Herramienta de Auditoría de Pasivos
"""

from dataclasses import dataclass
from typing import List, Any, Optional, Tuple
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


@dataclass
class TablaAmortizacion:
    """Resultado de una tabla de amortización."""
    concepto: str
    metodo: str
    encabezados: List[str]
    filas: List[List[Any]]
    info_extra: List[str]
    total_intereses: float
    total_pagado: float


def calcular_amortizacion_capital_fijo(
    concepto: str,
    monto_total: float,
    tasa_periodica: float,
    total_pagos: int,
    periodos_gracia: int = 0
) -> TablaAmortizacion:
    """
    Calcula tabla de amortización con método de Capital Fijo.
    """
    encabezados = ['Periodo', 'S. Inicial', 'Interés', 'Amortización', 'Pago', 'S. Final']
    filas = []
    total_intereses = 0.0

    # Periodo 0
    filas.append([0, 0.0, 0.0, 0.0, 0.0, round(monto_total, 2)])

    saldo = monto_total
    amortizacion = monto_total / total_pagos

    # Periodos de gracia
    for g in range(1, periodos_gracia + 1):
        filas.append([g, round(saldo, 2), 0.0, 0.0, 0.0, round(saldo, 2)])

    # Pagos
    for p in range(1, total_pagos + 1):
        interes = saldo * tasa_periodica
        pago = interes + amortizacion
        fin = saldo - amortizacion

        if p == total_pagos:
            fin = 0.0

        filas.append([
            p + periodos_gracia,
            round(saldo, 2),
            round(interes, 2),
            round(amortizacion, 2),
            round(pago, 2),
            round(fin, 2)
        ])

        total_intereses += interes
        saldo = fin

    return TablaAmortizacion(
        concepto=concepto,
        metodo="Capital Fijo",
        encabezados=encabezados,
        filas=filas,
        info_extra=[
            f"Monto Total: ${monto_total:,.2f}",
            f"Tasa Periódica: {tasa_periodica:.4%}",
            f"Amortización Fija: ${amortizacion:,.2f}"
        ],
        total_intereses=round(total_intereses, 2),
        total_pagado=round(monto_total + total_intereses, 2)
    )


def calcular_amortizacion_vencimiento(
    concepto: str,
    monto_total: float,
    tasa_periodica: float,
    total_pagos: int,
    periodos_gracia: int = 0
) -> TablaAmortizacion:
    """
    Calcula tabla de amortización con método de Pago al Vencimiento.
    """
    encabezados = ['Periodo', 'S. Inicial', 'Interés', 'Pago', 'S. Final']
    filas = []
    total_intereses = 0.0

    # Periodo 0
    filas.append([0, 0.0, 0.0, 0.0, round(monto_total, 2)])

    saldo = monto_total
    suma_intereses = 0.0

    # Periodos de gracia
    for g in range(1, periodos_gracia + 1):
        filas.append([g, round(saldo, 2), 0.0, 0.0, round(saldo, 2)])

    # Pagos
    for p in range(1, total_pagos + 1):
        interes = saldo * tasa_periodica

        if p == total_pagos:
            pago = monto_total + suma_intereses + interes
            fin = 0.0
        else:
            pago = 0.0
            fin = saldo + interes

        filas.append([
            p + periodos_gracia,
            round(saldo, 2),
            round(interes, 2),
            round(pago, 2),
            round(fin, 2)
        ])

        suma_intereses += interes
        total_intereses += interes
        saldo = fin

    return TablaAmortizacion(
        concepto=concepto,
        metodo="Pago al Vencimiento",
        encabezados=encabezados,
        filas=filas,
        info_extra=[
            f"Monto Total: ${monto_total:,.2f}",
            f"Tasa Periódica: {tasa_periodica:.4%}",
            f"Pago Final Estimado: ${(monto_total + total_intereses):,.2f}"
        ],
        total_intereses=round(total_intereses, 2),
        total_pagado=round(monto_total + total_intereses, 2)
    )


def exportar_amortizacion_excel(tabla: TablaAmortizacion) -> bytes:
    """Genera archivo Excel con la tabla de amortización."""
    wb = Workbook()
    ws = wb.active

    # Estilos
    azul_relleno = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    blanco_letra = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    borde_fino = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    titulo_font = Font(name="Calibri", size=14, bold=True, color="1F4E79")
    normal_font = Font(name="Calibri", size=11)
    centro = Alignment(horizontal="center", vertical="center")
    derecha = Alignment(horizontal="right", vertical="center")

    fila_actual = 1

    # Título
    ws.merge_cells(start_row=fila_actual, start_column=1, end_row=fila_actual, end_column=len(tabla.encabezados))
    celda_titulo = ws.cell(row=fila_actual, column=1, value=f"SISTEMA DE AMORTIZACIÓN: {tabla.concepto} ({tabla.metodo})")
    celda_titulo.font = titulo_font
    celda_titulo.alignment = centro
    fila_actual += 1

    # Info extra
    for linea in tabla.info_extra:
        ws.merge_cells(start_row=fila_actual, start_column=1, end_row=fila_actual, end_column=len(tabla.encabezados))
        celda = ws.cell(row=fila_actual, column=1, value=linea)
        celda.font = Font(name="Calibri", size=11, italic=True, color="555555")
        celda.alignment = centro
        fila_actual += 1

    fila_actual += 1

    # Encabezados
    for col, encabezado in enumerate(tabla.encabezados, 1):
        celda = ws.cell(row=fila_actual, column=col, value=encabezado)
        celda.font = blanco_letra
        celda.fill = azul_relleno
        celda.alignment = centro
        celda.border = borde_fino
    fila_actual += 1

    # Datos
    for fila_datos in tabla.filas:
        for col, valor in enumerate(fila_datos, 1):
            celda = ws.cell(row=fila_actual, column=col)
            celda.font = normal_font
            celda.border = borde_fino

            if col == 1:
                celda.value = valor
                celda.alignment = centro
            else:
                celda.alignment = derecha
                celda.value = valor
                celda.number_format = '"$"#,##0.00'
        fila_actual += 1

    # Ajustar ancho de columnas
    for col in range(1, len(tabla.encabezados) + 1):
        max_len = max(len(str(tabla.encabezados[col - 1])), 12)
        ws.column_dimensions[get_column_letter(col)].width = max_len + 5

    # Guardar en bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()