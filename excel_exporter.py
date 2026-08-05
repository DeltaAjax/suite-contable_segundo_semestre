# excel_exporter.py
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def generar_excel_estados_financieros(esf, eri, flujo_indirecto, estado_cambios, empresa="Empresa Demo S.A.", periodo="Del 1 de enero al 31 de diciembre de 2025") -> bytes:
    wb = Workbook()
    
    # Estilos globales (Tema FACPYA - Guinda)
    font_titulo = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    font_subtitulo = Font(name="Calibri", size=11, italic=True, color="FFFFFF")
    font_seccion = Font(name="Calibri", size=11, bold=True, color="800000")
    font_bold = Font(name="Calibri", size=11, bold=True)
    font_normal = Font(name="Calibri", size=11)
    
    fill_header = PatternFill(start_color="800000", end_color="800000", fill_type="solid")
    fill_sub_header = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
    
    align_center = Alignment(horizontal="center", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    
    border_thin = Border(bottom=Side(style="thin", color="D3D3D3"))
    border_total = Border(top=Side(style="thin", color="000000"), bottom=Side(style="double", color="000000"))
    
    fmt_moneda = '"$"#,##0.00;("$"#,##0.00);"-"'

    def _encabezado(ws, titulo_reporte, num_cols=3):
        col_fin = get_column_letter(num_cols)
        ws.merge_cells(f"A1:{col_fin}1")
        ws.merge_cells(f"A2:{col_fin}2")
        ws.merge_cells(f"A3:{col_fin}3")
        
        ws["A1"] = empresa
        ws["A2"] = titulo_reporte
        ws["A3"] = periodo
        
        for r in range(1, 4):
            for col in range(1, num_cols + 1):
                cell = ws.cell(row=r, column=col)
                cell.fill = fill_header
                cell.alignment = align_center
                if r == 1:
                    cell.font = font_titulo
                elif r == 2:
                    cell.font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
                else:
                    cell.font = font_subtitulo
        ws.row_dimensions[1].height = 25
        ws.row_dimensions[2].height = 20
        ws.row_dimensions[3].height = 18

    # ----------------------------------------------------
    # TAB 1: Estado de Situación Financiera (ESF)
    # ----------------------------------------------------
    ws_esf = wb.active
    ws_esf.title = "ESF"
    _encabezado(ws_esf, "ESTADO DE SITUACIÓN FINANCIERA", 3)
    
    ws_esf.cell(row=5, column=1, value="Concepto").font = font_bold
    ws_esf.cell(row=5, column=2, value="Año Actual").font = font_bold
    ws_esf.cell(row=5, column=3, value="Año Anterior").font = font_bold
    for c in range(1, 4):
        ws_esf.cell(row=5, column=c).fill = fill_sub_header
        ws_esf.cell(row=5, column=c).alignment = align_center if c > 1 else align_left

    r_idx = 6
    def _escribir_bloque_esf(titulo, rubros):
        nonlocal r_idx
        ws_esf.cell(row=r_idx, column=1, value=titulo).font = font_seccion
        r_idx += 1
        inicio = r_idx
        for r in rubros:
            ws_esf.cell(row=r_idx, column=1, value=f"  {r.nif.etiqueta}").font = font_normal
            
            c_act = ws_esf.cell(row=r_idx, column=2, value=r.saldo_actual)
            c_act.number_format = fmt_moneda
            c_act.font = font_normal
            
            c_ant = ws_esf.cell(row=r_idx, column=3, value=r.saldo_anterior)
            c_ant.number_format = fmt_moneda
            c_ant.font = font_normal
            
            ws_esf.cell(row=r_idx, column=1).border = border_thin
            c_act.border = border_thin
            c_ant.border = border_thin
            r_idx += 1
        fin = r_idx - 1
        
        ws_esf.cell(row=r_idx, column=1, value=f"Total {titulo}").font = font_bold
        
        cell_tot_act = ws_esf.cell(row=r_idx, column=2, value=f"=SUM(B{inicio}:B{fin})" if fin >= inicio else 0)
        cell_tot_act.font = font_bold
        cell_tot_act.number_format = fmt_moneda
        cell_tot_act.border = border_total
        
        cell_tot_ant = ws_esf.cell(row=r_idx, column=3, value=f"=SUM(C{inicio}:C{fin})" if fin >= inicio else 0)
        cell_tot_ant.font = font_bold
        cell_tot_ant.number_format = fmt_moneda
        cell_tot_ant.border = border_total
        
        f_row = r_idx
        r_idx += 2
        return f_row

    r_act_circ = _escribir_bloque_esf("Activo Circulante", esf.activo_circulante)
    r_act_nocirc = _escribir_bloque_esf("Activo No Circulante", esf.activo_no_circulante)
    
    ws_esf.cell(row=r_idx, column=1, value="TOTAL ACTIVO").font = font_bold
    c_tot_act = ws_esf.cell(row=r_idx, column=2, value=f"=B{r_act_circ}+B{r_act_nocirc}")
    c_tot_act.font = font_bold
    c_tot_act.number_format = fmt_moneda
    c_tot_act.border = border_total
    
    c_tot_ant = ws_esf.cell(row=r_idx, column=3, value=f"=C{r_act_circ}+C{r_act_nocirc}")
    c_tot_ant.font = font_bold
    c_tot_ant.number_format = fmt_moneda
    c_tot_ant.border = border_total
    r_idx += 2

    r_pas_cp = _escribir_bloque_esf("Pasivo a Corto Plazo", esf.pasivo_corto_plazo)
    r_pas_lp = _escribir_bloque_esf("Pasivo a Largo Plazo", esf.pasivo_largo_plazo)
    
    ws_esf.cell(row=r_idx, column=1, value="TOTAL PASIVO").font = font_bold
    c_tot_p_act = ws_esf.cell(row=r_idx, column=2, value=f"=B{r_pas_cp}+B{r_pas_lp}")
    c_tot_p_act.font = font_bold
    c_tot_p_act.number_format = fmt_moneda
    c_tot_p_act.border = border_total
    
    c_tot_p_ant = ws_esf.cell(row=r_idx, column=3, value=f"=C{r_pas_cp}+C{r_pas_lp}")
    c_tot_p_ant.font = font_bold
    c_tot_p_ant.number_format = fmt_moneda
    c_tot_p_ant.border = border_total
    r_tot_pasivo = r_idx
    r_idx += 2

    r_cap_contrib = _escribir_bloque_esf("Capital Contribuido", esf.capital_contribuido)
    r_cap_ganado = _escribir_bloque_esf("Capital Ganado", esf.capital_ganado)
    
    ws_esf.cell(row=r_idx, column=1, value="TOTAL CAPITAL CONTABLE").font = font_bold
    c_tot_c_act = ws_esf.cell(row=r_idx, column=2, value=f"=B{r_cap_contrib}+B{r_cap_ganado}")
    c_tot_c_act.font = font_bold
    c_tot_c_act.number_format = fmt_moneda
    c_tot_c_act.border = border_total
    
    c_tot_c_ant = ws_esf.cell(row=r_idx, column=3, value=f"=C{r_cap_contrib}+C{r_cap_ganado}")
    c_tot_c_ant.font = font_bold
    c_tot_c_ant.number_format = fmt_moneda
    c_tot_c_ant.border = border_total
    r_tot_cap = r_idx
    r_idx += 2

    ws_esf.cell(row=r_idx, column=1, value="TOTAL PASIVO + CAPITAL").font = font_bold
    c_sum_act = ws_esf.cell(row=r_idx, column=2, value=f"=B{r_tot_pasivo}+B{r_tot_cap}")
    c_sum_act.font = font_bold
    c_sum_act.number_format = fmt_moneda
    c_sum_act.border = border_total
    
    c_sum_ant = ws_esf.cell(row=r_idx, column=3, value=f"=C{r_tot_pasivo}+C{r_tot_cap}")
    c_sum_ant.font = font_bold
    c_sum_ant.number_format = fmt_moneda
    c_sum_ant.border = border_total

    # ----------------------------------------------------
    # TAB 2: Estado de Resultados Integrales (ERI)
    # ----------------------------------------------------
    ws_eri = wb.create_sheet(title="ERI")
    _encabezado(ws_eri, "ESTADO DE RESULTADOS INTEGRAL", 2)
    
    filas_eri = [
        ("Ventas", eri.ventas),
        ("Costo de Ventas", eri.costo_ventas),
        ("Utilidad Bruta", eri.utilidad_bruta),
        ("Gastos de Venta", eri.gastos_venta),
        ("Gastos de Administración", eri.gastos_administracion),
        ("Gastos Generales", eri.gastos_generales),
        ("Utilidad antes de Otros", eri.utilidad_antes_otros),
        ("Otros Productos", eri.otros_productos),
        ("Otros Gastos", eri.otros_gastos),
        ("Neto Otros Productos/Gastos", eri.neto_otros_productos_gastos),
        ("Utilidad de Operación", eri.utilidad_operacion),
        ("Productos Financieros", eri.productos_financieros),
        ("Gastos Financieros", eri.gastos_financieros),
        ("RIF", eri.rif),
        ("Utilidad antes de Impuestos", eri.utilidad_antes_impuestos),
        ("ISR", eri.isr),
        ("PTU", eri.ptu),
        ("Impuestos a la Utilidad", eri.impuestos_utilidad),
        ("Utilidad Neta", eri.utilidad_neta),
        ("Otros Resultados Integrales (ORI)", eri.ori),
        ("Utilidad Integral", eri.utilidad_integral),
    ]
    
    ws_eri.cell(row=5, column=1, value="Concepto").font = font_bold
    ws_eri.cell(row=5, column=2, value="Monto").font = font_bold
    ws_eri.cell(row=5, column=1).fill = fill_sub_header
    ws_eri.cell(row=5, column=2).fill = fill_sub_header
    
    r_idx = 6
    for concepto, monto in filas_eri:
        es_res = "Utilidad" in concepto or "Neto" in concepto or "RIF" in concepto
        ws_eri.cell(row=r_idx, column=1, value=concepto).font = font_bold if es_res else font_normal
        c_monto = ws_eri.cell(row=r_idx, column=2, value=monto)
        c_monto.font = font_bold if es_res else font_normal
        c_monto.number_format = fmt_moneda
        c_monto.border = border_total if es_res else border_thin
        r_idx += 1

    # ----------------------------------------------------
    # TAB 3: Estado de Flujo de Efectivo (EFE)
    # ----------------------------------------------------
    ws_efe = wb.create_sheet(title="EFE")
    _encabezado(ws_efe, "ESTADO DE FLUJO DE EFECTIVO (MÉTODO INDIRECTO)", 2)
    
    ws_efe.cell(row=5, column=1, value="Concepto").font = font_bold
    ws_efe.cell(row=5, column=2, value="Monto").font = font_bold
    ws_efe.cell(row=5, column=1).fill = fill_sub_header
    ws_efe.cell(row=5, column=2).fill = fill_sub_header
    
    r_idx = 6
    def _escribir_seccion_efe(titulo, filas, total_monto):
        nonlocal r_idx
        ws_efe.cell(row=r_idx, column=1, value=titulo).font = font_seccion
        r_idx += 1
        inicio = r_idx
        for item in filas:
            ws_efe.cell(row=r_idx, column=1, value=f"  {item.concepto}").font = font_normal
            c = ws_efe.cell(row=r_idx, column=2, value=item.monto)
            c.font = font_normal
            c.number_format = fmt_moneda
            c.border = border_thin
            r_idx += 1
        fin = r_idx - 1
        
        ws_efe.cell(row=r_idx, column=1, value=f"Total {titulo}").font = font_bold
        c_tot = ws_efe.cell(row=r_idx, column=2, value=f"=SUM(B{inicio}:B{fin})" if fin >= inicio else 0)
        c_tot.font = font_bold
        c_tot.number_format = fmt_moneda
        c_tot.border = border_total
        r_tot = r_idx
        r_idx += 2
        return r_tot

    r_tot_op = _escribir_seccion_efe("Actividades de Operación", flujo_indirecto.filas_operacion, flujo_indirecto.total_operacion)
    r_tot_inv = _escribir_seccion_efe("Actividades de Inversión", flujo_indirecto.filas_inversion, flujo_indirecto.total_inversion)
    r_tot_fin = _escribir_seccion_efe("Actividades de Financiamiento", flujo_indirecto.filas_financiamiento, flujo_indirecto.total_financiamiento)
    
    # Incremento / Disminución neta de efectivo
    ws_efe.cell(row=r_idx, column=1, value="Incremento (Disminución) Neto de Efectivo").font = font_bold
    c_inc = ws_efe.cell(row=r_idx, column=2, value=f"=B{r_tot_op}+B{r_tot_inv}+B{r_tot_fin}")
    c_inc.font = font_bold
    c_inc.number_format = fmt_moneda
    c_inc.border = border_thin
    r_inc = r_idx
    r_idx += 1

    # Efectivo Inicial
    ws_efe.cell(row=r_idx, column=1, value="Efectivo y Equivalentes de Efectivo al Inicio del Periodo").font = font_normal
    c_ini = ws_efe.cell(row=r_idx, column=2, value=flujo_indirecto.efectivo_inicial)
    c_ini.font = font_normal
    c_ini.number_format = fmt_moneda
    c_ini.border = border_thin
    r_ini = r_idx
    r_idx += 1

    # Efectivo Final Calculado
    ws_efe.cell(row=r_idx, column=1, value="Efectivo y Equivalentes de Efectivo al Final del Periodo").font = font_bold
    c_fin = ws_efe.cell(row=r_idx, column=2, value=f"=B{r_inc}+B{r_ini}")
    c_fin.font = font_bold
    c_fin.number_format = fmt_moneda
    c_fin.border = border_total

    # ----------------------------------------------------
    # TAB 4: Estado de Cambios en el Capital Contable (ECCC)
    # ----------------------------------------------------
    ws_eccc = wb.create_sheet(title="ECCC")
    _encabezado(ws_eccc, "ESTADO DE CAMBIOS EN EL CAPITAL CONTABLE", 4)
    
    headers_eccc = ["Concepto", "Capital Contribuido", "Capital Ganado", "Total Capital Contable"]
    for c_idx, h in enumerate(headers_eccc, start=1):
        cell = ws_eccc.cell(row=5, column=c_idx, value=h)
        cell.font = font_bold
        cell.fill = fill_sub_header
        cell.alignment = align_center if c_idx > 1 else align_left

    r_idx = 6
    for fila in estado_cambios.filas:
        ws_eccc.cell(row=r_idx, column=1, value=fila.concepto).font = font_bold if fila.es_encabezado_categoria else font_normal
        
        c_cc = ws_eccc.cell(row=r_idx, column=2, value=fila.capital_contribuido if fila.capital_contribuido is not None else 0)
        c_cg = ws_eccc.cell(row=r_idx, column=3, value=fila.capital_ganado if fila.capital_ganado is not None else 0)
        c_tot = ws_eccc.cell(row=r_idx, column=4, value=f"=B{r_idx}+C{r_idx}")
        
        for c in (c_cc, c_cg, c_tot):
            c.font = font_bold if fila.es_encabezado_categoria else font_normal
            c.number_format = fmt_moneda
            c.border = border_total if fila.es_encabezado_categoria else border_thin
            
        r_idx += 1

    # Autoajuste universal de columnas para todas las pestañas
    for sheet in wb.worksheets:
        for col in sheet.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val = str(cell.value or '')
                if cell.coordinate in sheet.merged_cells:
                    continue
                if len(val) > max_len:
                    max_len = len(val)
            sheet.column_dimensions[col_letter].width = max(max_len + 4, 18)

    stream = io.BytesIO()
    wb.save(stream)
    return stream.getvalue()