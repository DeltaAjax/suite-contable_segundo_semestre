"""
depreciacion.py — Módulo de Cálculo de Depreciación NIF C-6
Herramienta de Auditoría de Activos Fijos
"""

from dataclasses import dataclass
from typing import List, Dict, Any, Optional, Tuple
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


@dataclass
class TablaDepreciacion:
    """Resultado de una tabla de depreciación."""
    concepto: str
    metodo: str
    encabezados: List[str]
    filas: List[List[Any]]
    info_extra: List[str]
    total_depreciado: float
    valor_libros_final: float


def calcular_linea_recta(
    concepto: str,
    valor_activo: float,
    valor_residual: float,
    vida_util: int
) -> TablaDepreciacion:
    """Calcula tabla de depreciación por método de Línea Recta."""
    depre_anual = (valor_activo - valor_residual) / vida_util
    saldo_inicial = valor_activo
    depre_acumulada = 0
    filas = []

    for año in range(1, vida_util + 1):
        v_libros = saldo_inicial - depre_anual
        depre_acumulada += depre_anual
        filas.append([
            año,
            round(saldo_inicial, 2),
            round(depre_anual, 2),
            round(depre_acumulada, 2),
            round(v_libros, 2)
        ])
        saldo_inicial = v_libros

    return TablaDepreciacion(
        concepto=concepto,
        metodo="Línea Recta",
        encabezados=['Año', 'Saldo inicial', 'Depreciación', 'Dep. acum.', 'Valor libros'],
        filas=filas,
        info_extra=[f"Costo: ${valor_activo:,.2f} | Residual: ${valor_residual:,.2f}"],
        total_depreciado=round(depre_acumulada, 2),
        valor_libros_final=round(v_libros, 2)
    )


def calcular_suma_digitos(
    concepto: str,
    valor_activo: float,
    valor_residual: float,
    vida_util: int
) -> TablaDepreciacion:
    """Calcula tabla de depreciación por método de Suma de Dígitos."""
    suma = sum(range(1, vida_util + 1))
    valor_a_depreciar = valor_activo - valor_residual
    depre_acumulada = 0
    saldo_inicial = valor_activo
    filas = []

    for i in range(1, vida_util + 1):
        num_factor = vida_util - i + 1
        str_factor = f"{num_factor}/{suma}"
        depre = valor_a_depreciar * (num_factor / suma)
        depre_acumulada += depre
        v_libros = valor_activo - depre_acumulada

        filas.append([
            i,
            round(saldo_inicial, 2),
            str_factor,
            round(depre, 2),
            round(depre_acumulada, 2),
            round(v_libros, 2)
        ])
        saldo_inicial = v_libros

    return TablaDepreciacion(
        concepto=concepto,
        metodo="Suma de Dígitos",
        encabezados=['Año', 'Saldo inicial', 'Factor', 'Depreciación', 'Dep. acum.', 'Valor libros'],
        filas=filas,
        info_extra=[f"Costo: ${valor_activo:,.2f} | Residual: ${valor_residual:,.2f}"],
        total_depreciado=round(depre_acumulada, 2),
        valor_libros_final=round(v_libros, 2)
    )


def calcular_unidades_produccion(
    concepto: str,
    valor_activo: float,
    valor_residual: float,
    vida_util: int,
    capacidad_total: float,
    tipo_unidad: str,
    uso_anual: List[float]
) -> TablaDepreciacion:
    """
    Calcula tabla de depreciación por método de Unidades de Producción.
    
    Args:
        uso_anual: Lista de unidades/KM usadas por año (debe tener longitud = vida_util)
    """
    if len(uso_anual) != vida_util:
        raise ValueError(f"Se requieren {vida_util} valores de uso anual")

    factor = (valor_activo - valor_residual) / capacidad_total
    depre_acumulada = 0
    filas = []

    for año in range(1, vida_util + 1):
        uso = uso_anual[año - 1]
        depre = uso * factor
        depre_acumulada += depre
        v_libros = max(valor_activo - depre_acumulada, valor_residual)

        filas.append([
            año,
            round(uso, 0),
            round(factor, 4),
            round(depre, 2),
            round(depre_acumulada, 2),
            round(v_libros, 2)
        ])

    return TablaDepreciacion(
        concepto=concepto,
        metodo=f"Unidades de Producción ({tipo_unidad})",
        encabezados=['Año', f'Unidades ({tipo_unidad})', 'Factor', 'Depreciación', 'Dep. acum.', 'Valor libros'],
        filas=filas,
        info_extra=[f"Factor por unidad: {factor:.4f}", f"Capacidad total: {capacidad_total:,.0f} {tipo_unidad}"],
        total_depreciado=round(depre_acumulada, 2),
        valor_libros_final=round(v_libros, 2)
    )


def calcular_saldos_decrecientes(
    concepto: str,
    valor_activo: float,
    valor_residual: float,
    vida_util: int
) -> TablaDepreciacion:
    """
    Calcula tabla de depreciación por método de Saldos Decrecientes.
    Incluye el candado NIF C-6: no bajar del valor residual.
    """
    factor = (1 / vida_util) * 2
    dep_acumulada = 0
    valor_en_libros = valor_activo
    filas = []

    for año in range(1, vida_util + 1):
        saldo_inicial = valor_en_libros

        if saldo_inicial <= valor_residual:
            dep_periodo = 0.0
        else:
            dep_teorica = saldo_inicial * factor
            # Candado NIF C-6: No bajar del valor residual
            if (saldo_inicial - dep_teorica) < valor_residual:
                dep_periodo = saldo_inicial - valor_residual
            else:
                dep_periodo = dep_teorica

        dep_acumulada += dep_periodo
        valor_en_libros = valor_activo - dep_acumulada
        if valor_en_libros < valor_residual:
            valor_en_libros = valor_residual

        filas.append([
            año,
            round(saldo_inicial, 2),
            factor,
            round(dep_periodo, 2),
            round(dep_acumulada, 2),
            round(valor_en_libros, 2)
        ])

    return TablaDepreciacion(
        concepto=concepto,
        metodo="Saldos Decrecientes (Doble Cuota)",
        encabezados=['Año', 'Saldo inicial', 'Factor', 'Depreciación', 'Dep. acum.', 'Valor libros'],
        filas=filas,
        info_extra=[f"Costo: ${valor_activo:,.2f} | Piso Residual: ${valor_residual:,.2f}", f"Factor: {factor*100:.0f}%"],
        total_depreciado=round(dep_acumulada, 2),
        valor_libros_final=round(valor_en_libros, 2)
    )


def exportar_depreciacion_excel(tabla: TablaDepreciacion) -> bytes:
    """Genera archivo Excel con la tabla de depreciación."""
    wb = Workbook()
    ws = wb.active

    # Estilos
    azul_relleno = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    blanco_letra = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    borde_fino = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    titulo_font = Font(name="Calibri", size=14, bold=True, color="1F4E79")
    normal_font = Font(name="Calibri", size=11)
    centro = Alignment(horizontal="center", vertical="center")
    derecha = Alignment(horizontal="right", vertical="center")

    fila_actual = 1

    # Título
    ws.merge_cells(start_row=fila_actual, start_column=1, end_row=fila_actual, end_column=len(tabla.encabezados))
    celda_titulo = ws.cell(row=fila_actual, column=1, value=f"SISTEMA DE DEPRECIACIÓN: {tabla.concepto} ({tabla.metodo})")
    celda_titulo.font = titulo_font
    celda_titulo.alignment = centro
    fila_actual += 1

    # Info extra
    for linea in tabla.info_extra:
        ws.merge_cells(start_row=fila_actual, start_column=1, end_row=fila_actual, end_column=len(tabla.encabezados))
        celda = ws.cell(row=fila_actual, column=1, value=linea)
        celda.font = Font(name="Calibri", size=11, italic=True, color="555555")
        celda.alignment = centro
        fila_actual += 1

    fila_actual += 1

    # Encabezados
    for col, encabezado in enumerate(tabla.encabezados, 1):
        celda = ws.cell(row=fila_actual, column=col, value=encabezado)
        celda.font = blanco_letra
        celda.fill = azul_relleno
        celda.alignment = centro
        celda.border = borde_fino
    fila_actual += 1

    # Datos
    for fila_datos in tabla.filas:
        for col, valor in enumerate(fila_datos, 1):
            celda = ws.cell(row=fila_actual, column=col)
            celda.font = normal_font
            celda.border = borde_fino

            if col == 1:
                celda.value = valor
                celda.alignment = centro
            else:
                celda.alignment = derecha
                encab = tabla.encabezados[col - 1].lower()

                if "factor" in encab:
                    celda.value = valor
                    if isinstance(valor, float):
                        if "saldos" in tabla.metodo.lower():
                            celda.number_format = '0%'
                        else:
                            celda.number_format = '0.000'
                    else:
                        celda.number_format = '@'
                elif "unidades" in encab or "km" in encab:
                    celda.value = valor
                    celda.number_format = '#,##0'
                else:
                    celda.value = valor
                    celda.number_format = '"$"#,##0.00'
        fila_actual += 1

    # Ajustar ancho de columnas
    for col in range(1, len(tabla.encabezados) + 1):
        max_len = max(len(str(tabla.encabezados[col - 1])), 12)
        ws.column_dimensions[get_column_letter(col)].width = max_len + 5

    # Guardar en bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()